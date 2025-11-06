# Load required base libraries
import glob
import os

# Load required installed libraries
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Import required datashop_toolbox libraries
from datashop_toolbox.odfhdr import OdfHeader


def generate_report(file_path: str, wildcard: str, outfile: str) -> None:
    """
    Generates a report based on the metadata from ODF files as an Excel file.

    Parameters
    ----------
    file_path : str
        The file path to where the ODF files are located.
    wildcard:
        The wildcard string to filter ODF files.
    outfile:
        The output file name.

    """

    report_headings = [
        'File Name',
        'File_Spec',
        'Country_Institute_Code',
        'Cruise_Number',
        'Organization',
        'Chief_Scientist',
        'Start_Date',
        'End_Date',
        'Platform',
        'Cruise_Name',
        'Cruise_Description',
        'Data_Type',
        'Event_Number',
        'Event_Qualifier1',
        'Event_Qualifier2',
        'Creation_Date',
        'Orig_Creation_Date',
        'Start_Date_Time',
        'End_Date_Time',
        'Initial_Latitude',
        'Initial_Longitude',
        'Min_Depth',
        'Max_Depth',
        'Sampling_Interval',
        'Sounding',
        'Depth_Off_Bottom',
        'Station_Name',
        'Set_Number',
        'Event_Comments',
        'Inst_Type',
        'Model',
        'Serial_Number',
        'Description'
    ]

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the default sheet (usually named 'Sheet')
    worksheet = workbook.active

    # Add the report headings as the first row
    if worksheet is None:
        raise ValueError("No active worksheet found in workbook.")
    
    worksheet.append(report_headings)

    os.chdir(file_path)
    odfFiles = glob.glob(wildcard)

    for odf_file in odfFiles:
        odf = OdfHeader()
        odf.read_odf(file_path + odf_file)
        meta = list()
        meta.append(odf_file)
        meta.append(odf.file_specification.strip("'"))
        meta.append(odf.cruise_header.country_institute_code)
        meta.append(odf.cruise_header.cruise_number.strip("'"))
        meta.append(odf.cruise_header.organization.strip("'"))
        meta.append(odf.cruise_header.chief_scientist.strip("'"))
        meta.append(odf.cruise_header.start_date.strip("'"))
        meta.append(odf.cruise_header.end_date.strip("'"))
        meta.append(odf.cruise_header.platform.strip("'"))
        meta.append(odf.cruise_header.cruise_name.strip("'"))
        meta.append(odf.cruise_header.cruise_description.strip("'"))
        meta.append(odf.event_header.data_type.strip("'"))
        meta.append(odf.event_header.event_number.strip("'"))
        meta.append(odf.event_header.event_qualifier1.strip("'"))
        meta.append(odf.event_header.event_qualifier2.strip("'"))
        meta.append(odf.event_header.creation_date.strip("'"))
        meta.append(odf.event_header.orig_creation_date.strip("'"))
        meta.append(odf.event_header.start_date_time.strip("'"))
        meta.append(odf.event_header.end_date_time.strip("'"))
        meta.append(odf.event_header.initial_latitude)
        meta.append(odf.event_header.initial_longitude)
        meta.append(odf.event_header.min_depth)
        meta.append(odf.event_header.max_depth)
        meta.append(odf.event_header.sampling_interval)
        meta.append(odf.event_header.sounding)
        meta.append(odf.event_header.depth_off_bottom)
        meta.append(odf.event_header.station_name.strip("'"))
        meta.append(odf.event_header.set_number.strip("'"))
        the_event_comments = ""
        for ec in odf.event_header.event_comments:
            the_event_comments = the_event_comments + " " + ec.strip("'")
        meta.append(the_event_comments)
        meta.append(odf.instrument_header.instrument_type.strip("'"))
        meta.append(odf.instrument_header.model.strip("'"))
        meta.append(odf.instrument_header.serial_number.strip("'"))
        meta.append(odf.instrument_header.description.strip("'"))

        # Add the metadata from the current ODF file to the report
        worksheet.append(meta)

        for i, col in enumerate(worksheet.columns, start=1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    max_length = max(len(str(cell.value)), max_length)
                finally:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column].width = adjusted_width

    # Center text horizontally and vertically
    for col, cname in enumerate(worksheet.columns):
        column_letter = get_column_letter(col + 1)
        for cell in worksheet[column_letter + ":" + column_letter]:
            cell.alignment = Alignment(horizontal='center')

    # Save the workbook to a file
    workbook.save(outfile)

    # Print a success message
    print("Excel file created successfully!")


if __name__ == "__main__":
    generate_report("C:/DFO-MPO/DEV/Data/2025/LAT2025146/CTD/DATASHOP_PROCESSING/Step_2_Apply_Calibrations/ODF/", "D*.odf", "CTD_Metadata.xlsx")